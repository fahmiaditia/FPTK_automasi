print('Importing Library.....')
from os import listdir
from os.path import join
import os
import openpyxl as opx
import time
from datetime import datetime


#===============================================================================
arr = os.getcwd()
ab = os.listdir('Sumber_SIMULASI')
tanggal_jam = datetime.now().strftime("Tanggal %d-%m-%Y _Pukul %H-%M-%S")
#===============================================================================
print("Opening Worksheet.....")
new_wr = opx.Workbook()
sheet = new_wr.active
time.sleep(3)
#===============================================================================





#===============================================================================
print("Making Worksheet.....")
list_header = [
    'ID', 'Nama Depan', 
    'Nama Belakang',
]
for x in range(len(list_header)):
    sheet.cell(row=1, column=x+1).value = list_header[x]
for i in range(len(list_header), 10):
    for ff in range(len(list_header), len(list_header)+5):
        sheet.cell(row=1, column=ff+1).value = "Benda_"+str(ff-2)
    for ff in range(len(list_header)+5, len(list_header)+10):
        sheet.cell(row=1, column=ff+1).value = "Harga_"+str(ff-7)
    for ff in range(len(list_header)+10, len(list_header)+15):
        sheet.cell(row=1, column=ff+1).value = "Keterangan_"+str(ff-12)


# for x in range()
time.sleep(3)
#===============================================================================





#===============================================================================
print("Append data to Worksheet.....")
time.sleep(3)

for x in range(len(ab)):
    list_data = []

    print("Diproses: ", x+1,".",ab[x])
    sumber_data = os.path.join(arr, 'Sumber_SIMULASI', ab[x])
    wb = opx.load_workbook(sumber_data, data_only=True)
    sheet_formulir_ptk = wb['Sheet_A']
    sheet_tabel_formulir = wb['Sheet_B']
    print(" "*100, end="\r")
 
   
    for a in range(len(list_header)):
        data = sheet_formulir_ptk.cell(row=a+1, column=2).value
        sheet.cell(row=x+2, column=a+1).value = data
    for b in range(1, 4):
        for c in range(1, 6):
            data_tabel = sheet_tabel_formulir.cell(row=c+1, column=b).value
            sheet.cell(row=x+2, column=c+3).value = data_tabel
            aa = sheet.cell(row=x+2, column=c+3).value
            list_data.append(aa)
    
    for i in range(len(list_data)):    
        sheet.cell(row=x+2, column=i+4).value = list_data[i]
    
    

#===============================================================================
penamaan = "hasil_SIMULASI_"+tanggal_jam+".xlsx"
print("tersimpan: ", penamaan)
new_wr.save(penamaan)

