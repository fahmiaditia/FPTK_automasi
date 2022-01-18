print('Import library....')
from os import listdir
from os.path import join
import os
import openpyxl as opx
from datetime import datetime

# mengakses directori file sumber
arr = os.getcwd()
ab = os.listdir('Sumber')
tanggal_jam = datetime.now().strftime("Tanggal %d-%m-%Y __Pukul %H-%M-%S")


new_wr = opx.Workbook()
sheet = new_wr.active

list_header = [
    "TANGGAL", "NAMA SEKOLAH", "NPSN", "ALAMAT SEKOLAH",
    "NAMA LENGKAP (TANPA GELAR)", "NIK / NO. PASSPORT (UNTUK WN)", "JENIS KELAMIN",
    "TEMPAT LAHIR", "TANGGAL LAHIR", "NAMA IBU KANDUNG", "ALAMAT JALAN",
    "RT", "RW", "NAMA DUSUN", "DESA / KELURAHAN", "KECAMATAN", "KODE POS", "AGAMA", "NPWP",
    "NAMA WAJIB PAJAK", "KEWARGANEGARAAN", "STATUS PERKAWINAN", "NAMA SUAMI / ISTRI",
    "NIP SUAMI / ISTRI", "PEKERJAAN SUAMI / ISTRI", "STATUS PERKAWINAN",
    "NIP", "NIY / NIGK", "NUPTK", "JENIS PTK", "SK PENGANGKATAN", "TMT PENGANGKAT",
    "LEMBAGA PENGANGKAT", "SK CPNS", "TMT PNS", "TMT PNS",
    "PANGKAT GOLONGAN", "SUMBER GAJI", "KARTU PEGAWAI",
    "KARTU ISTRI (KARIS) / KARTU SUAMI (KARSU)", "PUNYA LISENSI KEPALA SEKOLAH",
    "KEAHLIAN LABORATORIUM", "MAMPU MENANGANI KEBUTUHAN KHUSUS",
    "KEAHLIAN BRAILE", "KEAHLIAN BHS. ISYARAT",
    "NOMOR TELEPON RUMAH", "NOMOR HP", "EMAIL",
    "ID BANK", "NOMOR REKENING BANK", "REKENING ATAS NAMA",
    "NOMOR SURAT TUGAS", "TANGGAL SURAT TUGAS", "TMT TUGAS",
    "STATUS SEKOLAH UNTUK", "KELUAR KARENA", "TANGGAL KELUAR", "Jenis Sertifikasi_1", 
    "NOMOR SERTIFIKASI_1", "TAHUN SERTIFIKASI_1", "BIDANG STUDI_1", "NRG_1",
    "NOMOR SERTIFIKASI_1", "JENIS SERTIFIKASI_2","NOMOR SERTIFIKASI_2","TAHUN SERTIFIKASI_2",
    "BIDANG STUDI_2","NRG_2","NOMOR SERTIFIKASI_2","JENIS SERTIFIKASI_3","NOMOR SERTIFIKASI_3",
    "TAHUN SERTIFIKASI_3","BIDANG STUDI_3",
    "NRG_3",
    "NOMOR SERTIFIKASI_3",
    "JENIS SERTIFIKASI_4",
    "NOMOR SERTIFIKASI_4",
    "TAHUN SERTIFIKASI_4",
    "BIDANG STUDI_4",
    "NRG_4",
    "NOMOR SERTIFIKASI_4",
    "JENIS SERTIFIKASI_5",
    "NOMOR SERTIFIKASI_5",
    "TAHUN SERTIFIKASI_5",
    "BIDANG STUDI_5",
    "NRG_5",
    "NOMOR SERTIFIKASI_5",
    "JENIS SERTIFIKASI_6",
    "NOMOR SERTIFIKASI_6",
    "TAHUN SERTIFIKASI_6",
    "BIDANG STUDI_6",
    "NRG_6",
    "NOMOR SERTIFIKASI_6",
    "JENIS SERTIFIKASI_7",
    "NOMOR SERTIFIKASI_7",
    "TAHUN SERTIFIKASI_7",
    "BIDANG STUDI_7",
    "NRG_7",
    "NOMOR SERTIFIKASI_7",
    "BIDANG STUDI_A",
    "JENJANG PENDIDIKAN_A",
    "GELAR AKADEMIK_A",
    "SATUAN PENDIDIKAN FORMAL_A",
    "TAHUN MASUK_A",
    "TAHUN_LULUS_A",
    "NIM_A",
    "MATA_KULIAH_A",
    "SEMESTER_A",
    "IPK_A",
    "BIDANG STUDI_B",
    "JENJANG PENDIDIKAN_B",
    "GELAR AKADEMIK_B",
    "SATUAN PENDIDIKAN FORMAL_B",
    "TAHUN MASUK_B",
    "TAHUN_LULUS_B",
    "NIM_B",
    "MATA_KULIAH_B",
    "SEMESTER_B",
    "IPK_B",
    "BIDANG STUDI_C",
    "JENJANG PENDIDIKAN_C",
    "GELAR AKADEMIK_C",
    "SATUAN PENDIDIKAN FORMAL_C",
    "TAHUN MASUK_C",
    "TAHUN_LULUS_C",
    "NIM_C",
    "MATA_KULIAH_C",
    "SEMESTER_C",
    "IPK_C",
    "BIDANG STUDI_D",
    "JENJANG PENDIDIKAN_D",
    "GELAR AKADEMIK_D",
    "SATUAN PENDIDIKAN FORMAL_D",
    "TAHUN MASUK_D",
    "TAHUN_LULUS_D",
    "NIM_D",
    "MATA_KULIAH_D",
    "SEMESTER_D",
    "IPK_D",
    "BIDANG STUDI_E",
    "JENJANG PENDIDIKAN_E",
    "GELAR AKADEMIK_E",
    "SATUAN PENDIDIKAN FORMAL_E",
    "TAHUN MASUK_E",
    "TAHUN_LULUS_E",
    "NIM_E",
    "MATA_KULIAH_E",
    "SEMESTER_E",
    "IPK_E",
    "BIDANG STUDI_F",
    "JENJANG PENDIDIKAN_F",
    "GELAR AKADEMIK_F",
    "SATUAN PENDIDIKAN FORMAL_F",
    "TAHUN MASUK_F",
    "TAHUN_LULUS_F",
    "NIM_F",
    "MATA_KULIAH_F",
    "SEMESTER_F",
    "IPK_F",
    "BIDANG STUDI_G",
    "JENJANG PENDIDIKAN_G",
    "GELAR AKADEMIK_G",
    "SATUAN PENDIDIKAN FORMAL_G",
    "TAHUN MASUK_G",
    "TAHUN_LULUS_G",
    "NIM_G",
    "MATA_KULIAH_G",
    "SEMESTER_G",
    "IPK_G",
    "BIDANG STUDI_H",
    "JENJANG PENDIDIKAN_H",
    "GELAR AKADEMIK_H",
    "SATUAN PENDIDIKAN FORMAL_H",
    "TAHUN MASUK_H",
    "TAHUN_LULUS_H",
    "NIM_H",
    "MATA_KULIAH_H",
    "SEMESTER_H",
    "IPK_H",
    "BIDANG STUDI_I",
    "JENJANG PENDIDIKAN_I",
    "GELAR AKADEMIK_I",
    "SATUAN PENDIDIKAN FORMAL_I",
    "TAHUN MASUK_I",
    "TAHUN_LULUS_I",
    "NIM_I",
    "MATA_KULIAH_I",
    "SEMESTER_I",
    "IPK_I",
    "BIDANG STUDI_J",
    "JENJANG PENDIDIKAN_J",
    "GELAR AKADEMIK_J",
    "SATUAN PENDIDIKAN FORMAL_J",
    "TAHUN MASUK_J",
    "TAHUN_LULUS_J",
    "NIM_J",
    "MATA_KULIAH_J",
    "SEMESTER_J",
    "IPK_J",

]

print('Import loopping...')
for x in range(len(list_header)):
    sheet.cell(row=1, column=x+1).value = list_header[x]
    # sheet.cell(row=1, column=x+1).font = Font(bold=True, color='040CFF')


for x in range(len(ab)):
    sumber_data = os.path.join(arr, 'Sumber', ab[x])
    wb = opx.load_workbook(sumber_data, data_only=True)
    sheet_formulir_ptk = wb['Formulir_PTK']
    print(" "*100, end="\r")
 
   
    for a in range(len(list_header)):
        data = sheet_formulir_ptk.cell(row=a+1, column=6).value
        sheet.cell(row=x+2, column=a+1).value = data
        
        
    print("Diproses: ", x+1,".",ab[x])

penamaan = "hasil__"+tanggal_jam+".xlsx"
new_wr.save(penamaan)









