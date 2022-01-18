from os import listdir
from os.path import join
import os
import openpyxl as opx
from datetime import datetime


al = datetime.now().strftime("tanggal %d-%m-%Y jam %H-%M-%S")
print(al)




wb = opx.load_workbook('Sumber_SIMULASI/simulasi.xlsx', data_only=True)

sh = wb['Sheet_B']

print('jumlah column: ', sh.max_column)
print('jumlah row: ', sh.max_row)



for a in range(1, sh.max_column+1):
    print(sh.cell(row=1, column=a).value)