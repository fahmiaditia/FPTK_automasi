from os import listdir
from os.path import isfile, join
import os
import openpyxl as opx
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


# mengakses directori file sumber
arr = os.getcwd()
ab = os.listdir('Sumber')


workbook = opx.Workbook()
sheet = workbook.active

sheet.cell(row=1, column=1).value = 'Nama Lengkap'
sheet.cell(row=1, column=1).font = Font(bold=True)

sheet.cell(row=1, column=2).value = 'Email'
sheet.cell(row=1, column=2).font = Font(bold=True)


sheet.cell(row=1, column=3).value = 'Gender'
sheet.cell(row=1, column=3).font = Font(bold=True)

sheet.cell(row=1, column=4).value = 'IP Address'
sheet.cell(row=1, column=4).font = Font(bold=True)



for x in range(len(ab)):
    ac = os.path.join(arr,'Sumber',ab[x])
    ap = opx.load_workbook(ac)
    sheet_obj = ap.active
    
    
    nama_depan = sheet_obj.cell(row = 2, column = 3).value
    nama_belakang = sheet_obj.cell(row = 3, column = 3).value
    
    
    email = sheet_obj.cell(row = 4, column = 3).value
    
    gender = sheet_obj.cell(row = 5, column = 3).value
    
    ip_add = sheet_obj.cell(row = 6, column = 3).value
    
    sheet.cell(row=x+2, column=1).value = nama_depan + " " + nama_belakang
    sheet.cell(row=x+2, column=2).value = email
    sheet.cell(row=x+2, column=3).value = gender
    sheet.cell(row=x+2, column=4).value = ip_add
    
    print('Selesai: ', x+1)
    
    
    
workbook.save('hasil.xlsx')

