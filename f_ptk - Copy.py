from os import listdir
from os.path import join
import os
import openpyxl as opx



# mengakses directori file sumber
arr = os.getcwd()
ab = os.listdir('Sumber')

new_wr = opx.Workbook()
sheet = new_wr.active

list_header = [
    "Tanggal", "Nama Sekolah", "NPSN", "Alamat Sekolah",
    "Nama Lengkap (Tanpa Gelar)", "NIK / No. Passport (Untuk WN)", "Jenis Kelamin",
    "Tempat Lahir", "Tanggal Lahir", "Nama Ibu Kandung", "Alamat Jalan",
    "RT", "RW", "Nama Dusun", "Desa / Kelurahan", "Kecamatan", "Kode POS", "Agama", "NPWP",
    "Nama Wajib Pajak", "Kewarganegaraan", "Status Perkawinan", "Nama Suami / Istri",
    "NIP Suami / Istri", "Pekerjaan Suami / Istri", "Status Perkawinan",
    "NIP", "NIY / NIGK", "NUPTK", "Jenis PTK", "SK Pengangkatan", "TMT Pengangkat",
    "Lembaga Pengangkat", "SK CPNS", "TMT PNS", "TMT PNS",
    "Pangkat Golongan", "Sumber Gaji", "Kartu Pegawai",
    "Kartu Istri (KARIS) / Kartu Suami (KARSU)", "Punya Lisensi Kepala Sekolah",
    "Keahlian Laboratorium", "Mampu Menangani Kebutuhan Khusus",
    "Keahlian Braile", "Keahlian Bhs. Isyarat",
    "Nomor telepon rumah", "Nomor HP", "Email",
    "Id Bank", "Nomor Rekening Bank", "Rekening Atas Nama",
    "Nomor Surat Tugas", "Tanggal Surat Tugas", "TMT Tugas",
    "Status Sekolah Untuk", "Keluar Karena", "Tanggal Keluar","FILE ASAL"
]

for x in range(len(list_header)):
    sheet.cell(row=1, column=x+1).value = list_header[x]
    # sheet.cell(row=1, column=x+1).font = Font(bold=True, color='040CFF')


for x in range(len(ab)):
    sumber_data = os.path.join(arr, 'Sumber', ab[x])
    wb = opx.load_workbook(sumber_data, data_only=True)
    sheet_formulir_ptk = wb['Formulir_PTK']
    print(" "*100, end="\r")
 
   
    for a in range(len(list_header)):
        data = sheet_formulir_ptk.cell(row=a+1, column=6).value
        sheet.cell(row=x+2, column=a+1).value = data
        sheet.cell(row=x+2, column=58).value = ab[x]
        
        
    print("Diproses: ", x+1,".",ab[x])

new_wr.save('hasil.xlsx')









